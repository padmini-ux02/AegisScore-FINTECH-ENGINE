import io
import csv
from datetime import datetime
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ReportLab imports for PDF
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

class ReportGenerator:
    @staticmethod
    def generate_pdf(data: Dict[str, Any]) -> io.BytesIO:
        """
        Generates a premium, publication-grade PDF credit assessment report.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=54,
            leftMargin=54,
            topMargin=54,
            bottomMargin=54
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles for a clean premium look
        primary_color = colors.HexColor("#0f172a") # Slate 900
        secondary_color = colors.HexColor("#475569") # Slate 600
        accent_color = colors.HexColor("#0284c7") # Sky 600
        border_color = colors.HexColor("#cbd5e1") # Slate 300
        bg_light = colors.HexColor("#f8fafc") # Slate 50
        
        # Set up custom paragraph styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=24,
            leading=28,
            textColor=primary_color,
            spaceAfter=15
        )
        
        subtitle_style = ParagraphStyle(
            'DocSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=secondary_color,
            spaceAfter=20
        )
        
        h2_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=18,
            textColor=primary_color,
            spaceBefore=15,
            spaceAfter=8,
            keepWithNext=True
        )
        
        body_style = ParagraphStyle(
            'BodyTextCustom',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#1e293b"),
            spaceAfter=6
        )
        
        bold_body_style = ParagraphStyle(
            'BoldBodyTextCustom',
            parent=body_style,
            fontName='Helvetica-Bold'
        )
        
        value_style = ParagraphStyle(
            'ValueText',
            parent=body_style,
            alignment=2 # Right align
        )
        
        disclaimer_style = ParagraphStyle(
            'DisclaimerCustom',
            parent=styles['Normal'],
            fontName='Helvetica-Oblique',
            fontSize=8,
            leading=11,
            textColor=secondary_color,
            spaceBefore=20
        )

        story = []
        
        # --- HEADER SECTION ---
        story.append(Paragraph("AegisScore", title_style))
        report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        story.append(Paragraph(f"Credit Assessment & Financial Risk Report &bull; Generated: {report_date}", subtitle_style))
        story.append(Spacer(1, 10))
        
        # --- EXECUTIVE SUMMARY (GRID WIDGETS) ---
        # Draw a highlighted block containing scores
        score = data["credit_score"]
        risk_cat = data["risk_category"]
        app_prob = data["approval_probability"]
        def_prob = data["default_probability"]
        health = data["financial_health_score"]
        
        # Color based on risk category
        risk_colors = {
            "Excellent": "#16a34a", # Green
            "Good": "#22c55e",
            "Moderate": "#eab308", # Yellow
            "High Risk": "#f97316", # Orange
            "Very High Risk": "#dc2626" # Red
        }
        risk_color_hex = risk_colors.get(risk_cat, "#475569")
        risk_color = colors.HexColor(risk_color_hex)
        
        summary_table_data = [
            [
                Paragraph("<b>Credit Score</b>", bold_body_style),
                Paragraph("<b>Risk Category</b>", bold_body_style),
                Paragraph("<b>Approval Prob.</b>", bold_body_style),
                Paragraph("<b>Default Risk</b>", bold_body_style),
                Paragraph("<b>Health Index</b>", bold_body_style)
            ],
            [
                Paragraph(f"<font size=18 color='{accent_color.hexval()}'><b>{score}</b></font><br/>Scale: 300-850", body_style),
                Paragraph(f"<font size=14 color='{risk_color_hex}'><b>{risk_cat}</b></font>", body_style),
                Paragraph(f"<font size=16><b>{app_prob * 100:.1f}%</b></font>", body_style),
                Paragraph(f"<font size=16><b>{def_prob * 100:.1f}%</b></font>", body_style),
                Paragraph(f"<font size=16 color='#0ea5e9'><b>{health:.1f}/100</b></font>", body_style)
            ]
        ]
        
        summary_table = Table(summary_table_data, colWidths=[1.1*inch, 1.3*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), bg_light),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOX', (0,0), (-1,-1), 1.5, primary_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('TOPPADDING', (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # --- CUSTOMER & FINANCIAL DETAILS ---
        story.append(Paragraph("Profile Overview", h2_style))
        
        inputs = data["inputs"]
        
        # Profile Details Table (Left: Demographics, Right: Financials)
        details_data = [
            [Paragraph("<b>Demographics</b>", bold_body_style), "", Paragraph("<b>Financial Details</b>", bold_body_style), ""],
            [Paragraph("Age", body_style), Paragraph(str(inputs.age), value_style), Paragraph("Annual Income", body_style), Paragraph(f"${inputs.annual_income:,.2f}", value_style)],
            [Paragraph("Gender", body_style), Paragraph(inputs.gender, value_style), Paragraph("Monthly Income", body_style), Paragraph(f"${inputs.monthly_income:,.2f}", value_style)],
            [Paragraph("Marital Status", body_style), Paragraph(inputs.marital_status, value_style), Paragraph("Existing Loan Bal.", body_style), Paragraph(f"${inputs.existing_loans_balance:,.2f}", value_style)],
            [Paragraph("Employment Type", body_style), Paragraph(inputs.employment_type, value_style), Paragraph("Debt-to-Income (DTI)", body_style), Paragraph(f"{inputs.debt_to_income_ratio * 100:.1f}%", value_style)],
            [Paragraph("Emp. Duration", body_style), Paragraph(f"{inputs.employment_duration} yrs", value_style), Paragraph("Credit Util. (CUR)", body_style), Paragraph(f"{inputs.credit_utilization_ratio * 100:.1f}%", value_style)],
            [Paragraph("Education Level", body_style), Paragraph(inputs.education_level, value_style), Paragraph("Savings Balance", body_style), Paragraph(f"${inputs.savings_amount:,.2f}", value_style)],
            ["", "", Paragraph("Investments", body_style), Paragraph(f"${inputs.investments:,.2f}", value_style)],
        ]
        
        details_table = Table(details_data, colWidths=[1.8*inch, 1.1*inch, 1.8*inch, 1.3*inch])
        details_table.setStyle(TableStyle([
            ('SPAN', (0,0), (1,0)),
            ('SPAN', (2,0), (3,0)),
            ('BACKGROUND', (0,0), (-1,0), bg_light),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LINEBELOW', (0,0), (-1,0), 1.0, primary_color),
            ('LINEBELOW', (0,1), (-1,-1), 0.5, border_color),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        
        story.append(details_table)
        story.append(Spacer(1, 15))
        
        # --- CREDIT HISTORY TABLE ---
        story.append(Paragraph("Credit History", h2_style))
        history_data = [
            [Paragraph("Metric", bold_body_style), Paragraph("Value", bold_body_style), Paragraph("Status / Evaluation", bold_body_style)],
            [
                Paragraph("Missed Payments (Last 2 Years)", body_style),
                Paragraph(str(inputs.missed_payments), value_style),
                Paragraph("Optimal" if inputs.missed_payments == 0 else "Needs Review", bold_body_style)
            ],
            [
                Paragraph("Active Loans", body_style),
                Paragraph(str(inputs.number_of_active_loans), value_style),
                Paragraph("Optimal" if inputs.number_of_active_loans <= 2 else "Elevated leverage", body_style)
            ],
            [
                Paragraph("Previous Loan Records", body_style),
                Paragraph(str(inputs.previous_loan_records), value_style),
                Paragraph(f"{inputs.previous_loan_records} settled accounts", body_style)
            ],
            [
                Paragraph("Payment History Ratio", body_style),
                Paragraph(f"{inputs.payment_history_ratio * 100:.1f}%", value_style),
                Paragraph("Excellent" if inputs.payment_history_ratio >= 0.95 else "Suboptimal", bold_body_style)
            ]
        ]
        history_table = Table(history_data, colWidths=[2.5*inch, 1.2*inch, 2.3*inch])
        history_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), bg_light),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW', (0,0), (-1,0), 1.0, primary_color),
            ('LINEBELOW', (0,1), (-1,-1), 0.5, border_color),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(history_table)
        story.append(Spacer(1, 15))
        
        # --- EXPLAINABLE AI SECTION ---
        story.append(Paragraph("Explainable AI (XAI) Insights", h2_style))
        
        xai = data["explainable_ai"]
        
        # Split explanations into columns
        rejection_reasons = xai.rejection_reasons if hasattr(xai, 'rejection_reasons') else xai.get('rejection_reasons', [])
        approval_reasons = xai.approval_reasons if hasattr(xai, 'approval_reasons') else xai.get('approval_reasons', [])
        
        rejections_str = "<br/>".join([f"&bull; {r}" for r in rejection_reasons])
        approvals_str = "<br/>".join([f"&bull; {a}" for a in approval_reasons])
        
        xai_data = [
            [Paragraph("<b>Approval Factors</b>", bold_body_style), Paragraph("<b>Risk / Mitigation Factors</b>", bold_body_style)],
            [Paragraph(approvals_str, body_style), Paragraph(rejections_str, body_style)]
        ]
        
        xai_table = Table(xai_data, colWidths=[3.0*inch, 3.0*inch])
        xai_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), bg_light),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOX', (0,0), (-1,-1), 1.0, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        
        story.append(xai_table)
        story.append(Spacer(1, 15))
        
        # --- RECOMMENDATIONS ---
        story.append(Paragraph("Actionable Financial Recommendations", h2_style))
        
        recommendations = []
        if inputs.missed_payments > 0:
            recommendations.append("<b>Establish Autopay:</b> Set up automatic drafts on all credit cards and loans to ensure zero missed payment recurrence.")
        if inputs.credit_utilization_ratio > 0.3:
            recommendations.append(f"<b>Pay Down Card Balances:</b> Your card utilization is at {inputs.credit_utilization_ratio*100:.1f}%. Make a mid-cycle payment to bring this below 30% to immediately boost your score.")
        if inputs.debt_to_income_ratio > 0.36:
            recommendations.append("<b>Deleverage Income:</b> Consolidate high-interest debts and focus on paying down active loan balances to lower your DTI below 36%.")
        if inputs.savings_amount < inputs.monthly_income * 3:
            recommendations.append("<b>Emergency Reserve:</b> Build liquid savings to match at least 3-6 months of expenses, improving your financial health index.")
        
        if not recommendations:
            recommendations.append("<b>Maintain Habits:</b> Excellent standing! Continue keeping utilization under 10% and automating payments to maintain score trajectory.")
            
        rec_str = "<br/><br/>".join([f"{i+1}. {r}" for i, r in enumerate(recommendations)])
        story.append(Paragraph(rec_str, body_style))
        
        # --- DISCLAIMER ---
        story.append(Spacer(1, 20))
        story.append(Paragraph(
            "<b>Disclaimer:</b> This credit score and assessment report is generated via machine learning models based on "
            "synthetic historical banking risk profiles. It is intended for simulation and analytical purposes only. "
            "AegisScore does not issue official loans, nor is it a licensed Credit Bureau.",
            disclaimer_style
        ))
        
        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_excel(data: Dict[str, Any]) -> io.BytesIO:
        """
        Generates a professionally formatted Excel spreadsheet of the credit report.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Credit Assessment"
        
        # Colors
        blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Navy Blue
        gray_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Light Gray
        white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)
        title_font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
        
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        
        # Title
        ws["A1"] = "AegisScore - Assessment Report"
        ws["A1"].font = title_font
        ws.merge_cells("A1:D1")
        ws.row_dimensions[1].height = 25
        
        # Date
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Calibri", size=9, italic=True)
        ws.merge_cells("A2:D2")
        
        # Section 1: Executive Summary
        ws["A4"] = "EXECUTIVE SUMMARY"
        ws["A4"].font = bold_font
        ws.merge_cells("A4:D4")
        
        summary_headers = ["Credit Score", "Risk Category", "Approval Prob.", "Default Risk", "Health Score"]
        summary_vals = [
            data["credit_score"],
            data["risk_category"],
            f"{data['approval_probability'] * 100:.2f}%",
            f"{data['default_probability'] * 100:.2f}%",
            data["financial_health_score"]
        ]
        
        for col_idx, (h, v) in enumerate(zip(summary_headers, summary_vals), start=1):
            cell_h = ws.cell(row=5, column=col_idx, value=h)
            cell_h.font = white_font
            cell_h.fill = blue_fill
            cell_h.alignment = Alignment(horizontal="center")
            cell_h.border = thin_border
            
            cell_v = ws.cell(row=6, column=col_idx, value=v)
            cell_v.font = regular_font
            cell_v.alignment = Alignment(horizontal="center")
            cell_v.border = thin_border
            
        # Section 2: Details
        ws["A8"] = "DEMOGRAPHICS"
        ws["A8"].font = bold_font
        ws["C8"] = "FINANCIAL DETAILS"
        ws["C8"].font = bold_font
        
        inputs = data["inputs"]
        
        demo_map = [
            ("Age", inputs.age),
            ("Gender", inputs.gender),
            ("Marital Status", inputs.marital_status),
            ("Employment Type", inputs.employment_type),
            ("Employment Duration", f"{inputs.employment_duration} years"),
            ("Education Level", inputs.education_level),
        ]
        
        fin_map = [
            ("Annual Income", inputs.annual_income),
            ("Monthly Income", inputs.monthly_income),
            ("Existing Loan Balances", inputs.existing_loans_balance),
            ("Debt-to-Income (DTI)", f"{inputs.debt_to_income_ratio * 100:.2f}%"),
            ("Credit Utilization (CUR)", f"{inputs.credit_utilization_ratio * 100:.2f}%"),
            ("Savings Balance", inputs.savings_amount),
            ("Investments", inputs.investments),
        ]
        
        # Fill Demographics (cols A-B)
        for i, (k, v) in enumerate(demo_map):
            row_idx = 9 + i
            ck = ws.cell(row=row_idx, column=1, value=k)
            ck.font = bold_font
            ck.border = thin_border
            cv = ws.cell(row=row_idx, column=2, value=v)
            cv.font = regular_font
            cv.border = thin_border
            
        # Fill Financials (cols C-D)
        for i, (k, v) in enumerate(fin_map):
            row_idx = 9 + i
            ck = ws.cell(row=row_idx, column=3, value=k)
            ck.font = bold_font
            ck.border = thin_border
            cv = ws.cell(row=row_idx, column=4, value=v)
            cv.font = regular_font
            cv.border = thin_border
            
            # format currency cells
            if k in ["Annual Income", "Monthly Income", "Existing Loan Balances", "Savings Balance", "Investments"]:
                cv.number_format = '$#,##0.00'
                
        # Section 3: Credit History (rows 17+)
        start_row = 17
        ws.cell(row=start_row, column=1, value="CREDIT HISTORY").font = bold_font
        
        hist_headers = ["Metric", "Value"]
        for col_idx, h in enumerate(hist_headers, start=1):
            cell = ws.cell(row=start_row+1, column=col_idx, value=h)
            cell.font = white_font
            cell.fill = blue_fill
            cell.border = thin_border
            
        hist_map = [
            ("Missed Payments (Last 2 Years)", inputs.missed_payments),
            ("Active Loans count", inputs.number_of_active_loans),
            ("Previous Loan Records", inputs.previous_loan_records),
            ("Payment History Ratio", f"{inputs.payment_history_ratio * 100:.2f}%"),
        ]
        
        for i, (k, v) in enumerate(hist_map):
            r = start_row + 2 + i
            ck = ws.cell(row=r, column=1, value=k)
            ck.font = regular_font
            ck.border = thin_border
            cv = ws.cell(row=r, column=2, value=v)
            cv.font = regular_font
            cv.border = thin_border
            
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and '$' in cell.number_format:
                    val_str = f"${val_str}"
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_csv(data: Dict[str, Any]) -> io.BytesIO:
        """
        Generates a standard flat CSV export of the record data.
        """
        buffer = io.BytesIO()
        # CSV requires text mode, so we wrap it
        wrapper = io.TextIOWrapper(buffer, encoding='utf-8', write_through=True)
        writer = csv.writer(wrapper)
        
        inputs = data["inputs"]
        
        # Headers
        headers = [
            "Generated At", "Credit Score", "Risk Category", "Approval Probability", 
            "Default Probability", "Financial Health Score", "Age", "Gender", 
            "Marital Status", "Employment Type", "Employment Duration", "Education Level",
            "Annual Income", "Monthly Income", "Existing Loan Balance", "Debt-to-Income Ratio",
            "Credit Utilization Ratio", "Savings Amount", "Investments", "Active Loans",
            "Previous Loan Records", "Missed Payments", "Payment History Ratio"
        ]
        
        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            data["credit_score"],
            data["risk_category"],
            f"{data['approval_probability']:.4f}",
            f"{data['default_probability']:.4f}",
            f"{data['financial_health_score']:.2f}",
            inputs.age,
            inputs.gender,
            inputs.marital_status,
            inputs.employment_type,
            inputs.employment_duration,
            inputs.education_level,
            inputs.annual_income,
            inputs.monthly_income,
            inputs.existing_loans_balance,
            inputs.debt_to_income_ratio,
            inputs.credit_utilization_ratio,
            inputs.savings_amount,
            inputs.investments,
            inputs.number_of_active_loans,
            inputs.previous_loan_records,
            inputs.missed_payments,
            inputs.payment_history_ratio
        ]
        
        writer.writerow(headers)
        writer.writerow(row)
        
        wrapper.detach() # Detach wrapper to return raw bytes buffer
        buffer.seek(0)
        return buffer
